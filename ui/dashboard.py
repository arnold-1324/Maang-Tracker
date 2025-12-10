# ui/dashboard.py
import sys
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv('.env.local')
load_dotenv() # Fallback to .env

# Add parent directory to path so we can import sibling packages
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from flask import Flask, render_template_string, request, redirect, url_for, jsonify, send_from_directory
from memory.db import (
    init_db, create_user, authenticate_user, get_user_by_id,
    save_user_credentials, get_user_credentials, CacheManager,
    get_weaknesses, get_all_job_postings
)
import subprocess
import threading
import time
import json
import asyncio
from datetime import datetime, timedelta
from flask_cors import CORS
from maang_agent.agent import get_mentor
from functools import wraps

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Initialize database
init_db()

# Cleanup expired cache on startup
CacheManager.cleanup_expired()

# Import services
from services.auth_service import generate_token, verify_token, get_user_from_token
from services.sync_service import SyncService

# Authentication decorator
def require_auth(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({"error": "No authorization token provided"}), 401
        
        if token.startswith('Bearer '):
            token = token[7:]
        
        user_id = get_user_from_token(token)
        if not user_id:
            return jsonify({"error": "Invalid or expired token"}), 401
        
        # Add user_id to request context
        request.user_id = user_id
        return f(*args, **kwargs)
    
    return decorated_function

# Try to register training blueprint if available
# try:
#     from dashboard.app.training.routes import training_bp
#     app.register_blueprint(training_bp)
#     print("[OK] Training blueprint registered successfully")
# except Exception as e:
#     # If the module doesn't exist yet, just log a warning
#     print(f"[WARN] Warning: Training routes not available: {e}")

# Try to register interview blueprint if available
socketio_instance = None
try:
    from ui.interview_routes import interview_bp, init_socketio
    app.register_blueprint(interview_bp)
    print("[OK] Interview blueprint registered successfully")
    socketio_instance = init_socketio(app)
    if socketio_instance:
        print("[OK] SocketIO initialized successfully")
    else:
        print("[WARN] Warning: SocketIO initialization returned None")
except Exception as e:
    # Interview module not available, continue without it
    import traceback
    print(f"[WARN] Warning: Interview routes not available: {e}")
    traceback.print_exc()

from tracker.tracker import snapshot_github, snapshot_leetcode, call_mcp
from roadmap.generator import recommend as get_recommendations

# ===== HTML TEMPLATES =====

MAIN_LAYOUT = """
<!doctype html>
<html>
<head>
    <title>MAANG Mentor Dashboard</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: #333;
            min-height: 100vh;
            padding: 20px;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }
        header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }
        header h1 { font-size: 2.5em; margin-bottom: 10px; }
        nav {
            display: flex;
            justify-content: center;
            gap: 20px;
            padding: 20px;
            background: #f8f9fa;
            border-bottom: 2px solid #667eea;
            flex-wrap: wrap;
        }
        nav a {
            text-decoration: none;
            color: #667eea;
            font-weight: bold;
            padding: 10px 20px;
            border-radius: 5px;
            transition: all 0.3s;
        }
        nav a:hover {
            background: #667eea;
            color: white;
            transform: translateY(-2px);
        }
        nav a.active {
            background: #667eea;
            color: white;
        }
        .content {
            padding: 30px;
        }
        .section {
            margin-bottom: 30px;
        }
        h2 {
            color: #667eea;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid #667eea;
        }
        .card {
            background: #f8f9fa;
            border-left: 4px solid #667eea;
            padding: 20px;
            margin: 10px 0;
            border-radius: 5px;
        }
        .btn {
            display: inline-block;
            background: #667eea;
            color: white;
            padding: 12px 30px;
            border-radius: 5px;
            text-decoration: none;
            transition: all 0.3s;
            border: none;
            cursor: pointer;
            font-size: 1em;
        }
        .btn:hover {
            background: #764ba2;
            transform: translateY(-2px);
            box-shadow: 0 5px 20px rgba(0,0,0,0.2);
        }
        .btn-secondary {
            background: #764ba2;
        }
        input, textarea {
            width: 100%;
            padding: 10px;
            margin: 10px 0;
            border: 1px solid #ddd;
            border-radius: 5px;
            font-family: inherit;
        }
        input:focus, textarea:focus {
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 5px rgba(102, 126, 234, 0.3);
        }
        .grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
        }
        .stats {
            text-align: center;
            padding: 20px;
            background: white;
            border-radius: 10px;
            border: 2px solid #667eea;
        }
        .stats .number {
            font-size: 2.5em;
            color: #667eea;
            font-weight: bold;
        }
        .stats .label {
            color: #666;
            margin-top: 10px;
        }
        ul { margin-left: 20px; }
        li { margin: 10px 0; }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>🎯 MAANG Mentor Dashboard</h1>
            <p>Interview Preparation Platform</p>
        </header>
        
        <nav>
            <a href="/" class="nav-link">Home</a>
            <a href="/interview" class="nav-link">Interview</a>
            <a href="/roadmap" class="nav-link">Roadmap</a>
            <a href="/training" class="nav-link">Training</a>
            <a href="/weakness" class="nav-link">Weaknesses</a>
        </nav>
        
        <div class="content">
            {% block content %}{% endblock %}
        </div>
    </div>
</body>
</html>
"""

HOME_TEMPLATE = """
<div class="section">
    <h2>Welcome to MAANG Interview Prep</h2>
    <div style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; padding: 20px; border-radius: 10px; margin-bottom: 20px; text-align: center;">
        <h3 style="margin-bottom: 10px;">🎯 Target: March 2026</h3>
        <div style="font-size: 1.5em; font-weight: bold;" id="countdown-display">{{ days_remaining }} days remaining</div>
    </div>
    <div class="grid">
        <div class="stats">
            <div class="number">{{ total_problems }}</div>
            <div class="label">Total Problems</div>
        </div>
        <div class="stats">
            <div class="number">{{ completed_problems }}</div>
            <div class="label">Completed</div>
        </div>
        <div class="stats">
            <div class="number">{{ success_rate }}%</div>
            <div class="label">Success Rate</div>
        </div>
    </div>
</div>
<script>
    function updateCountdown() {
        const targetDate = new Date('2026-03-15');
        const now = new Date();
        const diff = targetDate - now;
        const days = Math.floor(diff / (1000 * 60 * 60 * 24));
        const hours = Math.floor((diff % (1000 * 60 * 60 * 24)) / (1000 * 60 * 60));
        const minutes = Math.floor((diff % (1000 * 60 * 60)) / (1000 * 60));
        document.getElementById('countdown-display').textContent = 
            days + ' days, ' + hours + ' hours, ' + minutes + ' minutes remaining';
    }
    updateCountdown();
    setInterval(updateCountdown, 60000);
</script>

<div class="section">
    <h2>⚡ Quick Actions</h2>
    <div class="grid">
        <div class="card">
            <h3>Start Interview</h3>
            <p>Practice coding interviews with real-time feedback</p>
            <a href="/interview" class="btn">Start Now</a>
        </div>
        <div class="card">
            <h3>View Roadmap</h3>
            <p>Check your learning path and progress</p>
            <a href="/roadmap" class="btn">View Roadmap</a>
        </div>
        <div class="card">
            <h3>Sync Progress</h3>
            <p>Update your LeetCode and GitHub stats</p>
            <form method="post" action="/sync" style="margin: 0;">
                GitHub: <input type="text" name="github" placeholder="username" value="{{ github }}" style="width: auto; display: inline; padding: 5px;"><br>
                LeetCode: <input type="text" name="leetcode" placeholder="username" value="{{ leetcode }}" style="width: auto; display: inline; padding: 5px;"><br>
                <button type="submit" class="btn">Sync Now</button>
            </form>
        </div>
    </div>
</div>

<div class="section">
    <h2>📊 Weakness Profile</h2>
    {% if weaknesses %}
        {% for w in weaknesses %}
            <div class="card">
                <strong>{{ w.topic }}</strong> — Score: {{ w.score }}/10
                <br><small>Last seen: {{ w.last_seen }}</small>
            </div>
        {% endfor %}
    {% else %}
        <p>No weaknesses recorded yet. Start practicing to build your profile!</p>
    {% endif %}
</div>
"""

INTERVIEW_TEMPLATE = """
<div class="section">
    <h2>🎤 Interview Simulation</h2>
    <div class="grid">
        <div class="card">
            <h3>Coding Interview</h3>
            <p>Practice LeetCode-style problems with live code execution</p>
            <div style="margin-top: 10px;">
                <select id="difficulty" style="width: 100%; padding: 8px; margin: 5px 0;">
                    <option>Select Difficulty</option>
                    <option value="easy">Easy</option>
                    <option value="medium">Medium</option>
                    <option value="hard">Hard</option>
                </select>
                <button class="btn" onclick="startCodingInterview()">Start Coding Interview</button>
            </div>
        </div>
        
        <div class="card">
            <h3>System Design</h3>
            <p>Discuss architecture and system design problems</p>
            <div style="margin-top: 10px;">
                <input type="text" id="design_topic" placeholder="Topic (e.g., URL shortener, Cache)">
                <button class="btn" onclick="startDesignInterview()">Start Design Interview</button>
            </div>
        </div>
        
        <div class="card">
            <h3>Behavioral</h3>
            <p>Practice STAR method and behavioral questions</p>
            <button class="btn" onclick="startBehavioralInterview()">Start Behavioral Interview</button>
        </div>
    </div>
</div>

<div class="section">
    <h2>📈 Recent Interviews</h2>
    <div id="recent-interviews">
        <p>Loading...</p>
    </div>
</div>

<script>
function startCodingInterview() {
    const difficulty = document.getElementById('difficulty').value;
    if (!difficulty || difficulty === 'Select Difficulty') {
        alert('Please select a difficulty');
        return;
    }
    fetch('/api/interview/problems/' + difficulty)
        .then(r => r.json())
        .then(d => {
            if (d.success && d.problems.length) {
                alert('Interview started! Problem: ' + d.problems[0].title);
                // In production, open interview editor
            }
        })
        .catch(e => alert('Error: ' + e));
}

function startDesignInterview() {
    const topic = document.getElementById('design_topic').value;
    if (!topic) {
        alert('Please enter a topic');
        return;
    }
    fetch('/api/interview/system-design/' + topic)
        .then(r => r.json())
        .then(d => {
            if (d.success) {
                alert('System Design Interview started!');
            }
        })
        .catch(e => alert('Error: ' + e));
}

function startBehavioralInterview() {
    fetch('/api/interview/behavioral-question')
        .then(r => r.json())
        .then(d => {
            if (d.success) {
                alert('Question: ' + d.question.question);
            }
        })
        .catch(e => alert('Error: ' + e));
}
</script>
"""

ROADMAP_TEMPLATE = """
<div class="section">
    <h2>📚 Learning Roadmap</h2>
    <p>Follow this structured path to master MAANG interview topics</p>
    
    {% if roadmap %}
        {% for rec in roadmap %}
            <div class="card">
                <strong style="color: #667eea; font-size: 1.2em;">{{ rec.topic }}</strong>
                <br>Score: {{ rec.score }}/10
                <ul>
                {% for r in rec.recommendations %}
                    <li>{{ r }}</li>
                {% endfor %}
                </ul>
            </div>
        {% endfor %}
    {% else %}
        <p>Sync your progress to see personalized recommendations.</p>
    {% endif %}
</div>

<div class="section">
    <h2>🎯 12-Week Sprint</h2>
    <div class="card">
        <h3>Week 1-2: Arrays & Strings</h3>
        <p>Master two-pointer, sliding window, and basic array operations</p>
        <progress style="width: 100%; height: 20px;" value="75" max="100"></progress>
    </div>
    <div class="card">
        <h3>Week 3-4: Trees & Graphs</h3>
        <p>Learn DFS, BFS, and tree traversals</p>
        <progress style="width: 100%; height: 20px;" value="50" max="100"></progress>
    </div>
    <div class="card">
        <h3>Week 5-8: Dynamic Programming</h3>
        <p>Solve DP problems and recognize patterns</p>
        <progress style="width: 100%; height: 20px;" value="25" max="100"></progress>
    </div>
    <div class="card">
        <h3>Week 9-10: System Design</h3>
        <p>Design scalable systems and databases</p>
        <progress style="width: 100%; height: 20px;" value="10" max="100"></progress>
    </div>
    <div class="card">
        <h3>Week 11-12: Behavioral & Review</h3>
        <p>Practice interviews and consolidate learning</p>
        <progress style="width: 100%; height: 20px;" value="5" max="100"></progress>
    </div>
</div>
"""

WEAKNESS_TEMPLATE = """
<div class="section">
    <h2>💪 Weakness Analysis</h2>
    <p>Topics where you need the most practice</p>
    
    {% if weaknesses %}
        <div class="grid">
        {% for w in weaknesses %}
            <div class="card">
                <strong style="color: #667eea;">{{ w.topic }}</strong>
                <br>Score: <span style="color: #ff6b6b; font-weight: bold;">{{ w.score }}/10</span>
                <br><small>Last attempted: {{ w.last_seen }}</small>
                <br>
                <button class="btn" style="width: 100%; margin-top: 10px; font-size: 0.9em;">Practice Now</button>
            </div>
        {% endfor %}
        </div>
    {% else %}
        <p>Complete some interviews to identify your weak areas!</p>
    {% endif %}
</div>

<div class="section">
    <h2>📈 Progress Overview</h2>
    <div class="card">
        <strong>Topics Completed:</strong> {{ completed_topics }}/{{ total_topics }}
        <progress style="width: 100%; height: 20px; margin-top: 10px;" value="{{ completed_topics }}" max="{{ total_topics }}"></progress>
    </div>
</div>
"""

# ===== ROUTES =====

@app.route("/", methods=["GET"])
def index():
    from datetime import datetime, timedelta
    github = request.args.get("github") or os.getenv("GITHUB_USERNAME", "")
    leetcode = request.args.get("leetcode") or os.getenv("LEETCODE_USERNAME", "")
    weaknesses = get_weaknesses()
    user_id = request.args.get("user_id", "default_user")
    
    # Calculate countdown to March 2026
    target_date = datetime(2026, 3, 15)
    now = datetime.now()
    days_remaining = (target_date - now).days
    
    # Get progress data if available
    try:
        from maang_agent.memory_persistence import get_memory_manager
        memory = get_memory_manager()
        summary = memory.get_user_summary(user_id)
        total_problems = summary.get('total_problems_solved', 0)
        completed_problems = total_problems
        success_rate = 78  # Could calculate from interview stats
    except:
        total_problems = 500
        completed_problems = 125
        success_rate = 78
    
    return render_template_string(
        MAIN_LAYOUT + HOME_TEMPLATE,
        github=github,
        leetcode=leetcode,
        weaknesses=weaknesses,
        total_problems=total_problems,
        completed_problems=completed_problems,
        success_rate=success_rate,
        days_remaining=days_remaining
    )

@app.route("/interview", methods=["GET"])
def interview():
    """Interview page - use template file if available"""
    from pathlib import Path
    interview_template_path = Path(__file__).parent / "templates" / "interview.html"
    if interview_template_path.exists():
        with open(interview_template_path, 'r', encoding='utf-8') as f:
            return f.read()
    return render_template_string(MAIN_LAYOUT + INTERVIEW_TEMPLATE)

@app.route("/roadmap", methods=["GET"])
def roadmap():
    from pathlib import Path
    roadmap_template_path = Path(__file__).parent / "templates" / "roadmap.html"
    if roadmap_template_path.exists():
        with open(roadmap_template_path, 'r', encoding='utf-8') as f:
            return f.read()
    recommendations = get_recommendations()
    return render_template_string(
        MAIN_LAYOUT + ROADMAP_TEMPLATE,
        roadmap=recommendations
    )

@app.route("/api/roadmap/visualization", methods=["GET"])
def roadmap_visualization():
    """API endpoint for roadmap visualization data"""
    from roadmap.enhanced_generator import get_roadmap_generator
    user_id = request.args.get('user_id', 'default_user')
    
    generator = get_roadmap_generator(user_id)
    visualization = generator.visualize_progress()
    progress_summary = generator._get_progress_summary()
    
    return jsonify({
        'success': True,
        'visualization': {
            **visualization,
            'progress_summary': progress_summary
        }
    })

@app.route("/weakness", methods=["GET"])
def weakness():
    weaknesses = get_weaknesses()
    return render_template_string(
        MAIN_LAYOUT + WEAKNESS_TEMPLATE,
        weaknesses=weaknesses,
        completed_topics=8,
        total_topics=19
    )

@app.route("/training", methods=["GET"])
def training():
    """Training page - use blueprint route if available, otherwise fallback"""
    # Check if training blueprint is registered
    try:
        # Try to redirect to training dashboard
        from flask import redirect
        return redirect('/training/dashboard')
    except:
        # Fallback to simple template
        html = """
<div class="section">
    <h2>🎓 Training Mode</h2>
    <p>Learn new concepts with guided lessons and practice problems</p>
    <div class="grid">
        <div class="card">
            <h3>Arrays & Strings</h3>
            <p>Master fundamental data structures</p>
            <a href="/training/dashboard" class="btn">Start Training</a>
        </div>
        <div class="card">
            <h3>Dynamic Programming</h3>
            <p>Learn DP patterns and optimizations</p>
            <a href="/training/dashboard" class="btn">Start Training</a>
        </div>
        <div class="card">
            <h3>System Design</h3>
            <p>Design large-scale systems</p>
            <a href="/training/dashboard" class="btn">Start Training</a>
        </div>
    </div>
</div>
"""
        return render_template_string(MAIN_LAYOUT + html)

@app.route("/sync", methods=["POST"])
def sync():
    github_user = request.form.get("github") or os.getenv("GITHUB_TOKEN")
    leetcode_user = request.form.get("leetcode") or os.getenv("LEETCODE_USERNAME")
    
    def _worker(g, l):
        try:
            if g:
                snapshot_github(g)
            if l:
                snapshot_leetcode(l)
        except:
            pass
    
    t = threading.Thread(target=_worker, args=(github_user, leetcode_user))
    t.daemon = True
    t.start()
    
    return redirect(url_for("index", github=github_user, leetcode=leetcode_user))

@app.route("/api/integration/progress", methods=["GET"])
def unified_progress():
    """Get unified progress across all modules"""
    user_id = request.args.get('user_id', 'default_user')
    try:
        from integration.main_pipeline import get_pipeline
        pipeline = get_pipeline(user_id)
        progress = pipeline.get_unified_progress()
        return jsonify({'success': True, 'progress': progress})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route("/api/integration/sync", methods=["POST"])
def sync_modules():
    """Force sync across all modules"""
    user_id = request.json.get('user_id', 'default_user') if request.json else 'default_user'
    try:
        from integration.main_pipeline import get_pipeline
        pipeline = get_pipeline(user_id)
        result = pipeline.sync_all_modules()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route("/analyze", methods=["POST"])
def analyze():
    content = request.json
    code = content.get("code", "")
    func = content.get("func", "")
    try:
        from analyzer.complexity_analyzer import static_analysis, micro_benchmark
        static = static_analysis(code)
        micro = {}
        if func and content.get("param_gen"):
            try:
                micro = micro_benchmark(code, func, content.get("param_gen"))
            except Exception as e:
                micro = {"error": str(e)}
        return jsonify({"static": static, "micro": micro})
    except:
        return jsonify({"error": "Analyzer not available"}), 400

@app.route("/api/chat", methods=["POST"])
def chat_api():
    """Chat API for Next.js frontend"""
    data = request.json
    user_id = data.get("user_id", "default_user")
    message = data.get("message", "")
    context = data.get("context", {})
    
    if not message:
        return jsonify({"error": "Message is required"}), 400
        
    try:
        mentor = get_mentor(user_id)
        if not mentor.current_session_id:
            mentor.start_session(
                session_id=f"session_{int(time.time())}", 
                session_type="chat", 
                context=context
            )
            
        response = mentor.process_user_input(message, context)
        return jsonify({
            "response": response,
            "session_id": mentor.current_session_id
        })
    except Exception as e:
        print(f"Chat error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

        return jsonify({"error": str(e)}), 500

@app.route("/api/roadmap", methods=["GET"])
def get_roadmap_api():
    """Get roadmap data for visualization"""
    user_id = request.args.get("user_id", "default_user")
    try:
        from roadmap.enhanced_generator import get_roadmap_generator
        generator = get_roadmap_generator(user_id)
        roadmap_data = generator.get_learning_roadmap(weeks=12)
        visualization = generator.visualize_progress()
        
        return jsonify({
            "success": True,
            "roadmap": roadmap_data,
            "visualization": visualization
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/roadmap/topic-details", methods=["GET"])
def get_topic_details_api():
    """Get solved problems and notes for a specific topic"""
    user_id = request.args.get("user_id", "default_user")
    topic = request.args.get("topic", "")
    
    try:
        from maang_agent.memory_persistence import get_memory_manager
        memory = get_memory_manager()
        
        # Get all mastery records
        mastery = memory.get_problem_mastery(user_id)
        
        topic_problems = []
        
        # Mock data/fallback if DB is empty for demo/testing
        if not mastery:
            if topic in ["Two Pointers", "Arrays & Hashing", "Arrays"]:
                topic_problems = [
                    {
                        "problem_id": "two-sum",
                        "problem_name": "Two Sum",
                        "solved": True,
                        "difficulty": "Easy",
                        "starred": True,
                        "code": "class Solution:\n    def twoSum(self, nums: List[int], target: int) -> List[int]:\n        seen = {}\n        for i, num in enumerate(nums):\n            diff = target - num\n            if diff in seen:\n                return [seen[diff], i]\n            seen[num] = i",
                        "notes": "Used a hash map to store complements. Time complexity O(n).",
                        "language": "python"
                    }
                ]
        else:
            # Real DB Filter logic
            for m in mastery:
                # Basic check - normally check category
                topic_problems.append({
                    "problem_id": m.get("problem_id"),
                    "problem_name": m.get("problem_name", m.get("problem_id")), 
                    "solved": True,
                    "difficulty": "Medium", # Placeholder if not stored
                    "starred": False,
                    "last_practiced": m.get("last_practiced")
                 })
                 
        return jsonify({
            "success": True, 
            "problems": topic_problems,
            "topic": topic
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/jobs", methods=["GET"])
def get_jobs_api():
    """Get job matches and status"""
    try:
        import json
        import os
        
        # 1. Get legacy JSON data
        data = { "application_data": { "jobs_shortlist": [] } }
        job_file_path = os.path.join(os.getcwd(), 'tracker', 'job_matches.json')
        
        if os.path.exists(job_file_path):
            with open(job_file_path, 'r') as f:
                json_data = json.load(f)
                if json_data:
                    data = json_data
        
        # 2. Get DB jobs
        db_jobs = get_all_job_postings()
        
        # 3. Merge (Simple append for now, could dedup by URL)
        current_list = data.get("application_data", {}).get("jobs_shortlist", [])
        
        # Add DB jobs if they aren't duplicates (check by URL or Company+Role)
        existing_urls = set(j.get('url') for j in current_list if j.get('url'))
        
        for job in db_jobs:
            if not job.get('url') or job.get('url') not in existing_urls:
                # Ensure structure matches UI expectations
                current_list.append({
                    "id": job['id'], # Use DB ID
                    "company": job['company'],
                    "role": job['role'],
                    "location": job['location'] or "Remote",
                    "status": job['status'],
                    "notes": job['notes'] or "Crawled Job",
                    "url": job['url'],
                    "description": job['description'] # Crucial for ATS
                })
                
        data["application_data"]["jobs_shortlist"] = current_list

        return jsonify({"success": True, "data": data})
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/compiler/run", methods=["POST"])
def run_code_api():
    """Run code with custom input"""
    data = request.json
    code = data.get("code", "")
    language = data.get("language", "python")
    input_data = data.get("input", "")
    
    try:
        from interview.compiler import CodeCompiler
        compiler = CodeCompiler()
        result = compiler.compile_and_run(code, language, input_data, stdin_input=True)
        
        return jsonify({
            "success": result.success,
            "output": result.output,
            "error": result.error,
            "execution_time_ms": result.execution_time_ms
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/jobs/crawl", methods=["POST"])
def crawl_job_api():
    """Crawl job details from URL"""
    url = request.json.get("url")
    if not url:
        return jsonify({"success": False, "error": "URL required"}), 400
        
    try:
        from services.job_crawler import JobCrawler
        crawler = JobCrawler()
        result = crawler.fetch_job_details(url)
        
        if result["success"]:
            # Optional: Auto-save if requested
            if request.json.get("save", False):
                job_id = crawler.save_job(result["data"])
                result["data"]["id"] = job_id
        
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/jobs/auto-search", methods=["POST"])
def auto_search_jobs_api():
    """Automatically search for jobs based on resume keywords"""
    data = request.json
    keywords = data.get("keywords", "Software Engineer")
    location = data.get("location", "Remote")

    try:
        from services.job_crawler import JobCrawler
        crawler = JobCrawler()
        
        results = crawler.search_jobs(keywords, location)
        
        # Save all found jobs to DB automatically
        saved_jobs = []
        for job in results:
            job_id = crawler.save_job(job)
            if job_id:
                job['id'] = job_id
                saved_jobs.append(job)
            
        return jsonify({"success": True, "count": len(saved_jobs), "jobs": saved_jobs})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/jobs/auto-apply", methods=["POST"])
def auto_apply_job_api():
    """Apply for a job with tracking and email confirmation"""
    data = request.json
    job_id = data.get("job_id")
    user_email = data.get("user_email", "arnold.gnanaselvam@example.com")  # Default for demo
    
    if not job_id:
        return jsonify({"success": False, "error": "Job ID required"}), 400
        
    try:
        from services.application_tracker import ApplicationTracker
        from memory.db import get_conn
        
        # Get job details
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM job_postings WHERE id = ?", (job_id,))
        job_row = cur.fetchone()
        conn.close()
        
        if not job_row:
            return jsonify({"success": False, "error": "Job not found"}), 404
        
        job_data = {
            "id": job_row['id'],
            "company": job_row['company'],
            "role": job_row['title'],
            "location": job_row['location'] or "Remote",
            "url": job_row['url']
        }
        
        # Initialize tracker
        tracker = ApplicationTracker()
        
        # Record application
        app_record = tracker.record_application(
            job_data=job_data,
            resume_path="resumes/optimized_resume.pdf",  # Would be actual path
            ats_score=data.get("ats_score", 85.0)
        )
        
        if not app_record:
            return jsonify({"success": False, "error": "Failed to record application"}), 500
        
        # Send confirmation email
        email_sent = tracker.send_confirmation_email(
            user_email=user_email,
            job_data=job_data,
            tracking_id=app_record['tracking_id']
        )
        
        return jsonify({
            "success": True, 
            "status": "Applied",
            "message": f"Application submitted successfully. {'Confirmation email sent.' if email_sent else 'Email pending.'}",
            "tracking_id": app_record['tracking_id'],
            "application_id": app_record['application_id'],
            "applied_at": app_record['applied_at']
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/applications/status/<tracking_id>", methods=["GET"])
def get_application_status_api(tracking_id):
    """Get application status by tracking ID"""
    try:
        from services.application_tracker import ApplicationTracker
        
        tracker = ApplicationTracker()
        status = tracker.get_application_status(tracking_id)
        
        if status:
            return jsonify({"success": True, "data": status})
        else:
            return jsonify({"success": False, "error": "Application not found"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/applications", methods=["GET"])
def get_all_applications_api():
    """Get all applications for the user"""
    try:
        from services.application_tracker import ApplicationTracker
        
        tracker = ApplicationTracker()
        applications = tracker.get_all_applications(user_id=1)  # Default user
        
        return jsonify({"success": True, "data": applications})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/resume/generate-pdf", methods=["POST"])
def generate_resume_pdf_api():
    """Generate PDF from LaTeX resume content"""
    data = request.json
    latex_content = data.get("latex_content")
    job_title = data.get("job_title", "")
    
    if not latex_content:
        return jsonify({"success": False, "error": "LaTeX content required"}), 400
    
    try:
        from services.latex_pdf_converter import LatexToPdfConverter
        from datetime import datetime
        
        converter = LatexToPdfConverter()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        pdf_path = converter.generate_resume_pdf(
            latex_content=latex_content,
            job_title=job_title,
            timestamp=timestamp
        )
        
        if pdf_path:
            # Return file path for download
            return jsonify({
                "success": True,
                "pdf_path": pdf_path,
                "filename": os.path.basename(pdf_path),
                "message": "PDF generated successfully"
            })
        else:
            return jsonify({
                "success": False,
                "error": "PDF generation failed. Check server logs for details."
            }), 500
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/resume/download/<path:filename>", methods=["GET"])
def download_resume_pdf(filename):
    """Download generated PDF resume"""
    try:
        from flask import send_file
        pdf_dir = os.path.join(os.getcwd(), "resumes", "generated")
        file_path = os.path.join(pdf_dir, filename)
        
        if os.path.exists(file_path):
            return send_file(
                file_path,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
        else:
            return jsonify({"success": False, "error": "File not found"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/resume/analyze", methods=["POST"])
def analyze_resume_api():
    """Analyze resume text against job description"""
    data = request.json
    resume_text = data.get("resume_text", "")
    job_description = data.get("job_description", "")
    
    # If the resume text is the mock string, try to load the actual file content instead
    if "Experienced Backend Engineer" in resume_text or not resume_text:
        try:
             import os
             base_resume_path = os.path.join(os.getcwd(), 'resume_arnold_sde.tex')
             if os.path.exists(base_resume_path):
                 with open(base_resume_path, 'r', encoding='utf-8') as f:
                     resume_text = f.read()
        except Exception:
            pass # Fallback to provided text

    if not resume_text or not job_description:
        return jsonify({"success": False, "error": "Both resume text and job description are required"}), 400
        
    try:
        from services.resume_analyzer import Resumeanalyzer
        analyzer = Resumeanalyzer()
        analysis = analyzer.analyze_resume(resume_text, job_description)
        
        return jsonify({
            "success": True,
            "analysis": analysis
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/resume/optimize", methods=["POST"])
def optimize_resume_api():
    """Generate optimized resume LaTeX"""
    data = request.json
    missing_skills = data.get("missing_skills", [])
    job_description = data.get("job_description", "") # Needed for re-scoring check

    try:
        from services.resume_analyzer import Resumeanalyzer
        import os

        analyzer = Resumeanalyzer()

        # Load base resume
        base_resume_path = os.path.join(os.getcwd(), 'resume_arnold_sde.tex')
        if not os.path.exists(base_resume_path):
             return jsonify({"success": False, "error": "Base resume not found"}), 404

        with open(base_resume_path, 'r', encoding='utf-8') as f:
            original_latex = f.read()

        # Optimize
        optimized_latex = analyzer.optimize_resume_latex(original_latex, missing_skills)

        # Calculate Projected Score
        jd_keywords = analyzer._extract_skills(job_description.lower())
        total_skills = len(jd_keywords)

        original_analysis = analyzer.analyze_resume(original_latex, job_description)
        original_score = original_analysis['score']

        new_matches = original_analysis['matched_keywords'] + missing_skills
        new_matches = list(set(new_matches)) 

        if total_skills > 0:
            new_score = (len(new_matches) / total_skills) * 100
            new_score = min(100, new_score) 
            new_score = round(new_score, 1)
        else:
            new_score = 95.0 

        return jsonify({
            "success": True,
            "data": {
                "original_latex": original_latex,
                "optimized_latex": optimized_latex,
                "original_score": original_score,
                "new_score": new_score
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
@app.route("/api/progress", methods=["GET"])
def get_progress_api():
    """Get user progress analytics"""
    user_id = request.args.get("user_id", "default_user")
    
    try:
        from maang_agent.memory_persistence import get_memory_manager
        memory = get_memory_manager()
        
        mastery = memory.get_problem_mastery(user_id)
        solved_problems = len([m for m in mastery if m.get('optimal_solution_found')]) if mastery else 0
        total_problems = 350
        
        # Simple AI recommendations stub
        recommendations = [
            "Focus on Dynamic Programming",
            "Review Graph Algorithms",
            "Solve 2 problems daily"
        ]
        
        return jsonify({
            "success": True,
            "data": {
                "overall_mastery": min(100, (solved_problems / max(1, total_problems)) * 100),
                "problems_solved": solved_problems,
                "total_problems": total_problems,
                "recommendations": recommendations,
                "mock_interview_stats": {
                    "sessions_completed": 0,
                    "avg_score": 0
                }
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



@app.route("/api/roadmap/leetcode", methods=["GET"])
def get_leetcode_roadmap():
    """Get roadmap data directly from LeetCode via direct function call"""
    username = request.args.get("username") or os.getenv("LEETCODE_USERNAME") or "neal_wu"
    
    # Initialize Memory Manager
    from maang_agent.memory_persistence import AgentMemoryManager
    memory = AgentMemoryManager()
    
    # 1. Get User Stats & ALL Solved Problems (up to 1000 most recent)
    try:
        all_solved_res = leetcode_all_solved(username)
        print(f"Fetching all solved problems for: {username}")
    except Exception as e:
        print(f"Error fetching all solved problems: {e}")
        all_solved_res = {}

    solved_slugs = set()
    if all_solved_res and "data" in all_solved_res:
        submissions = all_solved_res["data"].get("recentAcSubmissionList", [])
        for sub in submissions:
            solved_slugs.add(sub["titleSlug"])
        print(f"Found {len(solved_slugs)} solved problems for {username}")
            
    # 2. Define Roadmap Topics
    topics = [
        {"id": "arrays", "name": "Arrays & Hashing", "tag": "array"},
        {"id": "two_pointers", "name": "Two Pointers", "tag": "two-pointers"},
        {"id": "stack", "name": "Stack", "tag": "stack"},
        {"id": "binary_search", "name": "Binary Search", "tag": "binary-search"},
        {"id": "sliding_window", "name": "Sliding Window", "tag": "sliding-window"},
        {"id": "linked_list", "name": "Linked List", "tag": "linked-list"},
        {"id": "trees", "name": "Trees", "tag": "tree"},
        {"id": "tries", "name": "Tries", "tag": "trie"},
        {"id": "backtracking", "name": "Backtracking", "tag": "backtracking"},
        {"id": "heap", "name": "Heap / Priority Queue", "tag": "heap"},
        {"id": "graphs", "name": "Graphs", "tag": "graph"},
        {"id": "dp_1d", "name": "1-D DP", "tag": "dynamic-programming"},
        {"id": "intervals", "name": "Intervals", "tag": "intervals"},
        {"id": "greedy", "name": "Greedy", "tag": "greedy"},
        {"id": "adv_graphs", "name": "Advanced Graphs", "tag": "adv_graphs"}, # Using 'adv_graphs' key in fallback
        {"id": "dp_2d", "name": "2-D DP", "tag": "dp_2d"}, # Using 'dp_2d' key in fallback
        {"id": "bit_manip", "name": "Bit Manipulation", "tag": "bit-manipulation"},
        {"id": "math", "name": "Math & Geometry", "tag": "math"}
    ]
    
    roadmap_data = []
    
    # Fallback data (Expanded to ~15 questions per topic for solid logic building)
    FALLBACK_PROBLEMS = {
        "array": [
            {"id": "1", "title": "Two Sum", "slug": "two-sum", "difficulty": "Easy"},
            {"id": "217", "title": "Contains Duplicate", "slug": "contains-duplicate", "difficulty": "Easy"},
            {"id": "242", "title": "Valid Anagram", "slug": "valid-anagram", "difficulty": "Easy"},
            {"id": "1929", "title": "Concatenation of Array", "slug": "concatenation-of-array", "difficulty": "Easy"},
            {"id": "1299", "title": "Replace Elements with Greatest Element on Right Side", "slug": "replace-elements-with-greatest-element-on-right-side", "difficulty": "Easy"},
            {"id": "392", "title": "Is Subsequence", "slug": "is-subsequence", "difficulty": "Easy"},
            {"id": "58", "title": "Length of Last Word", "slug": "length-of-last-word", "difficulty": "Easy"},
            {"id": "14", "title": "Longest Common Prefix", "slug": "longest-common-prefix", "difficulty": "Easy"},
            {"id": "49", "title": "Group Anagrams", "slug": "group-anagrams", "difficulty": "Medium"},
            {"id": "347", "title": "Top K Frequent Elements", "slug": "top-k-frequent-elements", "difficulty": "Medium"},
            {"id": "238", "title": "Product of Array Except Self", "slug": "product-of-array-except-self", "difficulty": "Medium"},
            {"id": "36", "title": "Valid Sudoku", "slug": "valid-sudoku", "difficulty": "Medium"},
            {"id": "128", "title": "Longest Consecutive Sequence", "slug": "longest-consecutive-sequence", "difficulty": "Medium"},
            {"id": "75", "title": "Sort Colors", "slug": "sort-colors", "difficulty": "Medium"},
            {"id": "53", "title": "Maximum Subarray", "slug": "maximum-subarray", "difficulty": "Medium"},
            {"id": "41", "title": "First Missing Positive", "slug": "first-missing-positive", "difficulty": "Hard"}
        ],
        "two-pointers": [
            {"id": "125", "title": "Valid Palindrome", "slug": "valid-palindrome", "difficulty": "Easy"},
            {"id": "680", "title": "Valid Palindrome II", "slug": "valid-palindrome-ii", "difficulty": "Easy"},
            {"id": "344", "title": "Reverse String", "slug": "reverse-string", "difficulty": "Easy"},
            {"id": "977", "title": "Squares of a Sorted Array", "slug": "squares-of-a-sorted-array", "difficulty": "Easy"},
            {"id": "283", "title": "Move Zeroes", "slug": "move-zeroes", "difficulty": "Easy"},
            {"id": "167", "title": "Two Sum II - Input Array Is Sorted", "slug": "two-sum-ii-input-array-is-sorted", "difficulty": "Medium"},
            {"id": "15", "title": "3Sum", "slug": "3sum", "difficulty": "Medium"},
            {"id": "11", "title": "Container With Most Water", "slug": "container-with-most-water", "difficulty": "Medium"},
            {"id": "189", "title": "Rotate Array", "slug": "rotate-array", "difficulty": "Medium"},
            {"id": "881", "title": "Boats to Save People", "slug": "boats-to-save-people", "difficulty": "Medium"},
            {"id": "18", "title": "4Sum", "slug": "4sum", "difficulty": "Medium"},
            {"id": "16", "title": "3Sum Closest", "slug": "3sum-closest", "difficulty": "Medium"},
            {"id": "80", "title": "Remove Duplicates from Sorted Array II", "slug": "remove-duplicates-from-sorted-array-ii", "difficulty": "Medium"},
            {"id": "1768", "title": "Merge Strings Alternately", "slug": "merge-strings-alternately", "difficulty": "Easy"},
            {"id": "42", "title": "Trapping Rain Water", "slug": "trapping-rain-water", "difficulty": "Hard"}
        ],
        "stack": [
            {"id": "20", "title": "Valid Parentheses", "slug": "valid-parentheses", "difficulty": "Easy"},
            {"id": "682", "title": "Baseball Game", "slug": "baseball-game", "difficulty": "Easy"},
            {"id": "1047", "title": "Remove All Adjacent Duplicates In String", "slug": "remove-all-adjacent-duplicates-in-string", "difficulty": "Easy"},
            {"id": "496", "title": "Next Greater Element I", "slug": "next-greater-element-i", "difficulty": "Easy"},
            {"id": "232", "title": "Implement Queue using Stacks", "slug": "implement-queue-using-stacks", "difficulty": "Easy"},
            {"id": "155", "title": "Min Stack", "slug": "min-stack", "difficulty": "Medium"},
            {"id": "150", "title": "Evaluate Reverse Polish Notation", "slug": "evaluate-reverse-polish-notation", "difficulty": "Medium"},
            {"id": "22", "title": "Generate Parentheses", "slug": "generate-parentheses", "difficulty": "Medium"},
            {"id": "739", "title": "Daily Temperatures", "slug": "daily-temperatures", "difficulty": "Medium"},
            {"id": "853", "title": "Car Fleet", "slug": "car-fleet", "difficulty": "Medium"},
            {"id": "901", "title": "Online Stock Span", "slug": "online-stock-span", "difficulty": "Medium"},
            {"id": "735", "title": "Asteroid Collision", "slug": "asteroid-collision", "difficulty": "Medium"},
            {"id": "394", "title": "Decode String", "slug": "decode-string", "difficulty": "Medium"},
            {"id": "71", "title": "Simplify Path", "slug": "simplify-path", "difficulty": "Medium"},
            {"id": "84", "title": "Largest Rectangle in Histogram", "slug": "largest-rectangle-in-histogram", "difficulty": "Hard"}
        ],
        "binary-search": [
            {"id": "704", "title": "Binary Search", "slug": "binary-search", "difficulty": "Easy"},
            {"id": "374", "title": "Guess Number Higher or Lower", "slug": "guess-number-higher-or-lower", "difficulty": "Easy"},
            {"id": "35", "title": "Search Insert Position", "slug": "search-insert-position", "difficulty": "Easy"},
            {"id": "278", "title": "First Bad Version", "slug": "first-bad-version", "difficulty": "Easy"},
            {"id": "69", "title": "Sqrt(x)", "slug": "sqrtx", "difficulty": "Easy"},
            {"id": "74", "title": "Search a 2D Matrix", "slug": "search-a-2d-matrix", "difficulty": "Medium"},
            {"id": "875", "title": "Koko Eating Bananas", "slug": "koko-eating-bananas", "difficulty": "Medium"},
            {"id": "153", "title": "Find Minimum in Rotated Sorted Array", "slug": "find-minimum-in-rotated-sorted-array", "difficulty": "Medium"},
            {"id": "33", "title": "Search in Rotated Sorted Array", "slug": "search-in-rotated-sorted-array", "difficulty": "Medium"},
            {"id": "981", "title": "Time Based Key-Value Store", "slug": "time-based-key-value-store", "difficulty": "Medium"},
            {"id": "162", "title": "Find Peak Element", "slug": "find-peak-element", "difficulty": "Medium"},
            {"id": "34", "title": "Find First and Last Position of Element in Sorted Array", "slug": "find-first-and-last-position-of-element-in-sorted-array", "difficulty": "Medium"},
            {"id": "744", "title": "Find Smallest Letter Greater Than Target", "slug": "find-smallest-letter-greater-than-target", "difficulty": "Easy"},
            {"id": "367", "title": "Valid Perfect Square", "slug": "valid-perfect-square", "difficulty": "Easy"},
            {"id": "4", "title": "Median of Two Sorted Arrays", "slug": "median-of-two-sorted-arrays", "difficulty": "Hard"}
        ],
        "sliding-window": [
            {"id": "121", "title": "Best Time to Buy and Sell Stock", "slug": "best-time-to-buy-and-sell-stock", "difficulty": "Easy"},
            {"id": "219", "title": "Contains Duplicate II", "slug": "contains-duplicate-ii", "difficulty": "Easy"},
            {"id": "1876", "title": "Substrings of Size Three with Distinct Characters", "slug": "substrings-of-size-three-with-distinct-characters", "difficulty": "Easy"},
            {"id": "643", "title": "Maximum Average Subarray I", "slug": "maximum-average-subarray-i", "difficulty": "Easy"},
            {"id": "3", "title": "Longest Substring Without Repeating Characters", "slug": "longest-substring-without-repeating-characters", "difficulty": "Medium"},
            {"id": "424", "title": "Longest Repeating Character Replacement", "slug": "longest-repeating-character-replacement", "difficulty": "Medium"},
            {"id": "567", "title": "Permutation in String", "slug": "permutation-in-string", "difficulty": "Medium"},
            {"id": "209", "title": "Minimum Size Subarray Sum", "slug": "minimum-size-subarray-sum", "difficulty": "Medium"},
            {"id": "904", "title": "Fruit Into Baskets", "slug": "fruit-into-baskets", "difficulty": "Medium"},
            {"id": "1004", "title": "Max Consecutive Ones III", "slug": "max-consecutive-ones-iii", "difficulty": "Medium"},
            {"id": "438", "title": "Find All Anagrams in a String", "slug": "find-all-anagrams-in-a-string", "difficulty": "Medium"},
            {"id": "1456", "title": "Maximum Number of Vowels in a Substring of Given Length", "slug": "maximum-number-of-vowels-in-a-substring-of-given-length", "difficulty": "Medium"},
            {"id": "1343", "title": "Number of Sub-arrays of Size K and Average Greater than or Equal to Threshold", "slug": "number-of-sub-arrays-of-size-k-and-average-greater-than-or-equal-to-threshold", "difficulty": "Medium"},
            {"id": "239", "title": "Sliding Window Maximum", "slug": "sliding-window-maximum", "difficulty": "Hard"},
            {"id": "76", "title": "Minimum Window Substring", "slug": "minimum-window-substring", "difficulty": "Hard"}
        ],
        "linked-list": [
            {"id": "206", "title": "Reverse Linked List", "slug": "reverse-linked-list", "difficulty": "Easy"},
            {"id": "21", "title": "Merge Two Sorted Lists", "slug": "merge-two-sorted-lists", "difficulty": "Easy"},
            {"id": "141", "title": "Linked List Cycle", "slug": "linked-list-cycle", "difficulty": "Easy"},
            {"id": "83", "title": "Remove Duplicates from Sorted List", "slug": "remove-duplicates-from-sorted-list", "difficulty": "Easy"},
            {"id": "234", "title": "Palindrome Linked List", "slug": "palindrome-linked-list", "difficulty": "Easy"},
            {"id": "203", "title": "Remove Linked List Elements", "slug": "remove-linked-list-elements", "difficulty": "Easy"},
            {"id": "160", "title": "Intersection of Two Linked Lists", "slug": "intersection-of-two-linked-lists", "difficulty": "Easy"},
            {"id": "876", "title": "Middle of the Linked List", "slug": "middle-of-the-linked-list", "difficulty": "Easy"},
            {"id": "143", "title": "Reorder List", "slug": "reorder-list", "difficulty": "Medium"},
            {"id": "19", "title": "Remove Nth Node From End of List", "slug": "remove-nth-node-from-end-of-list", "difficulty": "Medium"},
            {"id": "138", "title": "Copy List with Random Pointer", "slug": "copy-list-with-random-pointer", "difficulty": "Medium"},
            {"id": "2", "title": "Add Two Numbers", "slug": "add-two-numbers", "difficulty": "Medium"},
            {"id": "287", "title": "Find the Duplicate Number", "slug": "find-the-duplicate-number", "difficulty": "Medium"},
            {"id": "146", "title": "LRU Cache", "slug": "lru-cache", "difficulty": "Medium"},
            {"id": "23", "title": "Merge k Sorted Lists", "slug": "merge-k-sorted-lists", "difficulty": "Hard"}
        ],
        "tree": [
            {"id": "226", "title": "Invert Binary Tree", "slug": "invert-binary-tree", "difficulty": "Easy"},
            {"id": "104", "title": "Maximum Depth of Binary Tree", "slug": "maximum-depth-of-binary-tree", "difficulty": "Easy"},
            {"id": "543", "title": "Diameter of Binary Tree", "slug": "diameter-of-binary-tree", "difficulty": "Easy"},
            {"id": "110", "title": "Balanced Binary Tree", "slug": "balanced-binary-tree", "difficulty": "Easy"},
            {"id": "100", "title": "Same Tree", "slug": "same-tree", "difficulty": "Easy"},
            {"id": "572", "title": "Subtree of Another Tree", "slug": "subtree-of-another-tree", "difficulty": "Easy"},
            {"id": "235", "title": "Lowest Common Ancestor of a BST", "slug": "lowest-common-ancestor-of-a-binary-search-tree", "difficulty": "Medium"},
            {"id": "102", "title": "Binary Tree Level Order Traversal", "slug": "binary-tree-level-order-traversal", "difficulty": "Medium"},
            {"id": "199", "title": "Binary Tree Right Side View", "slug": "binary-tree-right-side-view", "difficulty": "Medium"},
            {"id": "1448", "title": "Count Good Nodes in Binary Tree", "slug": "count-good-nodes-in-binary-tree", "difficulty": "Medium"},
            {"id": "98", "title": "Validate Binary Search Tree", "slug": "validate-binary-search-tree", "difficulty": "Medium"},
            {"id": "230", "title": "Kth Smallest Element in a BST", "slug": "kth-smallest-element-in-a-bst", "difficulty": "Medium"},
            {"id": "105", "title": "Construct Binary Tree from Preorder and Inorder Traversal", "slug": "construct-binary-tree-from-preorder-and-inorder-traversal", "difficulty": "Medium"},
            {"id": "124", "title": "Binary Tree Maximum Path Sum", "slug": "binary-tree-maximum-path-sum", "difficulty": "Hard"},
            {"id": "297", "title": "Serialize and Deserialize Binary Tree", "slug": "serialize-and-deserialize-binary-tree", "difficulty": "Hard"}
        ],
        "trie": [
            {"id": "208", "title": "Implement Trie (Prefix Tree)", "slug": "implement-trie-prefix-tree", "difficulty": "Medium"},
            {"id": "211", "title": "Design Add and Search Words Data Structure", "slug": "design-add-and-search-words-data-structure", "difficulty": "Medium"},
            {"id": "212", "title": "Word Search II", "slug": "word-search-ii", "difficulty": "Hard"},
            {"id": "720", "title": "Longest Word in Dictionary", "slug": "longest-word-in-dictionary", "difficulty": "Medium"},
            {"id": "677", "title": "Map Sum Pairs", "slug": "map-sum-pairs", "difficulty": "Medium"},
            {"id": "648", "title": "Replace Words", "slug": "replace-words", "difficulty": "Medium"},
            {"id": "421", "title": "Maximum XOR of Two Numbers in an Array", "slug": "maximum-xor-of-two-numbers-in-an-array", "difficulty": "Medium"},
            {"id": "1268", "title": "Search Suggestions System", "slug": "search-suggestions-system", "difficulty": "Medium"},
            {"id": "1032", "title": "Stream of Characters", "slug": "stream-of-characters", "difficulty": "Hard"},
            {"id": "745", "title": "Prefix and Suffix Search", "slug": "prefix-and-suffix-search", "difficulty": "Hard"},
            {"id": "1707", "title": "Maximum XOR With an Element From Array", "slug": "maximum-xor-with-an-element-from-array", "difficulty": "Hard"},
            {"id": "336", "title": "Palindrome Pairs", "slug": "palindrome-pairs", "difficulty": "Hard"},
            {"id": "1804", "title": "Implement Trie II (Prefix Tree)", "slug": "implement-trie-ii-prefix-tree", "difficulty": "Medium"},
            {"id": "472", "title": "Concatenated Words", "slug": "concatenated-words", "difficulty": "Hard"},
            {"id": "139", "title": "Word Break", "slug": "word-break", "difficulty": "Medium"}
        ],
        "backtracking": [
            {"id": "78", "title": "Subsets", "slug": "subsets", "difficulty": "Medium"},
            {"id": "39", "title": "Combination Sum", "slug": "combination-sum", "difficulty": "Medium"},
            {"id": "46", "title": "Permutations", "slug": "permutations", "difficulty": "Medium"},
            {"id": "90", "title": "Subsets II", "slug": "subsets-ii", "difficulty": "Medium"},
            {"id": "40", "title": "Combination Sum II", "slug": "combination-sum-ii", "difficulty": "Medium"},
            {"id": "79", "title": "Word Search", "slug": "word-search", "difficulty": "Medium"},
            {"id": "131", "title": "Palindrome Partitioning", "slug": "palindrome-partitioning", "difficulty": "Medium"},
            {"id": "17", "title": "Letter Combinations of a Phone Number", "slug": "letter-combinations-of-a-phone-number", "difficulty": "Medium"},
            {"id": "51", "title": "N-Queens", "slug": "n-queens", "difficulty": "Hard"},
            {"id": "37", "title": "Sudoku Solver", "slug": "sudoku-solver", "difficulty": "Hard"},
            {"id": "77", "title": "Combinations", "slug": "combinations", "difficulty": "Medium"},
            {"id": "216", "title": "Combination Sum III", "slug": "combination-sum-iii", "difficulty": "Medium"},
            {"id": "47", "title": "Permutations II", "slug": "permutations-ii", "difficulty": "Medium"},
            {"id": "93", "title": "Restore IP Addresses", "slug": "restore-ip-addresses", "difficulty": "Medium"},
            {"id": "52", "title": "N-Queens II", "slug": "n-queens-ii", "difficulty": "Hard"}
        ],
        "heap": [
            {"id": "703", "title": "Kth Largest Element in a Stream", "slug": "kth-largest-element-in-a-stream", "difficulty": "Easy"},
            {"id": "1046", "title": "Last Stone Weight", "slug": "last-stone-weight", "difficulty": "Easy"},
            {"id": "973", "title": "K Closest Points to Origin", "slug": "k-closest-points-to-origin", "difficulty": "Medium"},
            {"id": "215", "title": "Kth Largest Element in an Array", "slug": "kth-largest-element-in-an-array", "difficulty": "Medium"},
            {"id": "621", "title": "Task Scheduler", "slug": "task-scheduler", "difficulty": "Medium"},
            {"id": "355", "title": "Design Twitter", "slug": "design-twitter", "difficulty": "Medium"},
            {"id": "295", "title": "Find Median from Data Stream", "slug": "find-median-from-data-stream", "difficulty": "Hard"},
            {"id": "23", "title": "Merge k Sorted Lists", "slug": "merge-k-sorted-lists", "difficulty": "Hard"},
            {"id": "347", "title": "Top K Frequent Elements", "slug": "top-k-frequent-elements", "difficulty": "Medium"},
            {"id": "378", "title": "Kth Smallest Element in a Sorted Matrix", "slug": "kth-smallest-element-in-a-sorted-matrix", "difficulty": "Medium"},
            {"id": "451", "title": "Sort Characters By Frequency", "slug": "sort-characters-by-frequency", "difficulty": "Medium"},
            {"id": "767", "title": "Reorganize String", "slug": "reorganize-string", "difficulty": "Medium"},
            {"id": "1337", "title": "The K Weakest Rows in a Matrix", "slug": "the-k-weakest-rows-in-a-matrix", "difficulty": "Easy"},
            {"id": "1834", "title": "Single-Threaded CPU", "slug": "single-threaded-cpu", "difficulty": "Medium"},
            {"id": "1985", "title": "Find the Kth Largest Integer in the Array", "slug": "find-the-kth-largest-integer-in-the-array", "difficulty": "Medium"}
        ],
        "intervals": [
            {"id": "57", "title": "Insert Interval", "slug": "insert-interval", "difficulty": "Medium"},
            {"id": "56", "title": "Merge Intervals", "slug": "merge-intervals", "difficulty": "Medium"},
            {"id": "435", "title": "Non-overlapping Intervals", "slug": "non-overlapping-intervals", "difficulty": "Medium"},
            {"id": "252", "title": "Meeting Rooms", "slug": "meeting-rooms", "difficulty": "Easy"},
            {"id": "253", "title": "Meeting Rooms II", "slug": "meeting-rooms-ii", "difficulty": "Medium"},
            {"id": "1851", "title": "Minimum Interval to Include Each Query", "slug": "minimum-interval-to-include-each-query", "difficulty": "Hard"},
            {"id": "228", "title": "Summary Ranges", "slug": "summary-ranges", "difficulty": "Easy"},
            {"id": "452", "title": "Minimum Number of Arrows to Burst Balloons", "slug": "minimum-number-of-arrows-to-burst-balloons", "difficulty": "Medium"},
            {"id": "1288", "title": "Remove Covered Intervals", "slug": "remove-covered-intervals", "difficulty": "Medium"},
            {"id": "986", "title": "Interval List Intersections", "slug": "interval-list-intersections", "difficulty": "Medium"},
            {"id": "729", "title": "My Calendar I", "slug": "my-calendar-i", "difficulty": "Medium"},
            {"id": "731", "title": "My Calendar II", "slug": "my-calendar-ii", "difficulty": "Medium"},
            {"id": "732", "title": "My Calendar III", "slug": "my-calendar-iii", "difficulty": "Hard"},
            {"id": "1272", "title": "Remove Interval", "slug": "remove-interval", "difficulty": "Medium"},
            {"id": "1353", "title": "Maximum Number of Events That Can Be Attended", "slug": "maximum-number-of-events-that-can-be-attended", "difficulty": "Medium"}
        ],
        "greedy": [
            {"id": "53", "title": "Maximum Subarray", "slug": "maximum-subarray", "difficulty": "Medium"},
            {"id": "55", "title": "Jump Game", "slug": "jump-game", "difficulty": "Medium"},
            {"id": "45", "title": "Jump Game II", "slug": "jump-game-ii", "difficulty": "Medium"},
            {"id": "134", "title": "Gas Station", "slug": "gas-station", "difficulty": "Medium"},
            {"id": "846", "title": "Hand of Straights", "slug": "hand-of-straights", "difficulty": "Medium"},
            {"id": "1899", "title": "Merge Triplets to Form Target Triplet", "slug": "merge-triplets-to-form-target-triplet", "difficulty": "Medium"},
            {"id": "763", "title": "Partition Labels", "slug": "partition-labels", "difficulty": "Medium"},
            {"id": "678", "title": "Valid Parenthesis String", "slug": "valid-parenthesis-string", "difficulty": "Medium"},
            {"id": "121", "title": "Best Time to Buy and Sell Stock", "slug": "best-time-to-buy-and-sell-stock", "difficulty": "Easy"},
            {"id": "122", "title": "Best Time to Buy and Sell Stock II", "slug": "best-time-to-buy-and-sell-stock-ii", "difficulty": "Medium"},
            {"id": "409", "title": "Longest Palindrome", "slug": "longest-palindrome", "difficulty": "Easy"},
            {"id": "179", "title": "Largest Number", "slug": "largest-number", "difficulty": "Medium"},
            {"id": "135", "title": "Candy", "slug": "candy", "difficulty": "Hard"},
            {"id": "334", "title": "Increasing Triplet Subsequence", "slug": "increasing-triplet-subsequence", "difficulty": "Medium"},
            {"id": "1871", "title": "Jump Game VII", "slug": "jump-game-vii", "difficulty": "Medium"}
        ],
        "adv_graphs": [
            {"id": "332", "title": "Reconstruct Itinerary", "slug": "reconstruct-itinerary", "difficulty": "Hard"},
            {"id": "1584", "title": "Min Cost to Connect All Points", "slug": "min-cost-to-connect-all-points", "difficulty": "Medium"},
            {"id": "743", "title": "Network Delay Time", "slug": "network-delay-time", "difficulty": "Medium"},
            {"id": "778", "title": "Swim in Rising Water", "slug": "swim-in-rising-water", "difficulty": "Hard"},
            {"id": "269", "title": "Alien Dictionary", "slug": "alien-dictionary", "difficulty": "Hard"},
            {"id": "787", "title": "Cheapest Flights Within K Stops", "slug": "cheapest-flights-within-k-stops", "difficulty": "Medium"},
            {"id": "1135", "title": "Connecting Cities With Minimum Cost", "slug": "connecting-cities-with-minimum-cost", "difficulty": "Medium"},
            {"id": "1192", "title": "Critical Connections in a Network", "slug": "critical-connections-in-a-network", "difficulty": "Hard"},
            {"id": "310", "title": "Minimum Height Trees", "slug": "minimum-height-trees", "difficulty": "Medium"},
            {"id": "210", "title": "Course Schedule II", "slug": "course-schedule-ii", "difficulty": "Medium"},
            {"id": "990", "title": "Satisfiability of Equality Equations", "slug": "satisfiability-of-equality-equations", "difficulty": "Medium"},
            {"id": "721", "title": "Accounts Merge", "slug": "accounts-merge", "difficulty": "Medium"},
            {"id": "802", "title": "Find Eventual Safe States", "slug": "find-eventual-safe-states", "difficulty": "Medium"},
            {"id": "127", "title": "Word Ladder", "slug": "word-ladder", "difficulty": "Hard"},
            {"id": "126", "title": "Word Ladder II", "slug": "word-ladder-ii", "difficulty": "Hard"}
        ],
        "dp_2d": [
            {"id": "62", "title": "Unique Paths", "slug": "unique-paths", "difficulty": "Medium"},
            {"id": "1143", "title": "Longest Common Subsequence", "slug": "longest-common-subsequence", "difficulty": "Medium"},
            {"id": "309", "title": "Best Time to Buy and Sell Stock with Cooldown", "slug": "best-time-to-buy-and-sell-stock-with-cooldown", "difficulty": "Medium"},
            {"id": "518", "title": "Coin Change 2", "slug": "coin-change-2", "difficulty": "Medium"},
            {"id": "494", "title": "Target Sum", "slug": "target-sum", "difficulty": "Medium"},
            {"id": "97", "title": "Interleaving String", "slug": "interleaving-string", "difficulty": "Medium"},
            {"id": "329", "title": "Longest Increasing Path in a Matrix", "slug": "longest-increasing-path-in-a-matrix", "difficulty": "Hard"},
            {"id": "115", "title": "Distinct Subsequences", "slug": "distinct-subsequences", "difficulty": "Hard"},
            {"id": "72", "title": "Edit Distance", "slug": "edit-distance", "difficulty": "Hard"},
            {"id": "312", "title": "Burst Balloons", "slug": "burst-balloons", "difficulty": "Hard"},
            {"id": "10", "title": "Regular Expression Matching", "slug": "regular-expression-matching", "difficulty": "Hard"},
            {"id": "63", "title": "Unique Paths II", "slug": "unique-paths-ii", "difficulty": "Medium"},
            {"id": "64", "title": "Minimum Path Sum", "slug": "minimum-path-sum", "difficulty": "Medium"},
            {"id": "120", "title": "Triangle", "slug": "triangle", "difficulty": "Medium"},
            {"id": "221", "title": "Maximal Square", "slug": "maximal-square", "difficulty": "Medium"}
        ],
        "bit-manipulation": [
            {"id": "136", "title": "Single Number", "slug": "single-number", "difficulty": "Easy"},
            {"id": "191", "title": "Number of 1 Bits", "slug": "number-of-1-bits", "difficulty": "Easy"},
            {"id": "338", "title": "Counting Bits", "slug": "counting-bits", "difficulty": "Easy"},
            {"id": "190", "title": "Reverse Bits", "slug": "reverse-bits", "difficulty": "Easy"},
            {"id": "268", "title": "Missing Number", "slug": "missing-number", "difficulty": "Easy"},
            {"id": "371", "title": "Sum of Two Integers", "slug": "sum-of-two-integers", "difficulty": "Medium"},
            {"id": "7", "title": "Reverse Integer", "slug": "reverse-integer", "difficulty": "Medium"},
            {"id": "231", "title": "Power of Two", "slug": "power-of-two", "difficulty": "Easy"},
            {"id": "342", "title": "Power of Four", "slug": "power-of-four", "difficulty": "Easy"},
            {"id": "1342", "title": "Number of Steps to Reduce a Number to Zero", "slug": "number-of-steps-to-reduce-a-number-to-zero", "difficulty": "Easy"},
            {"id": "461", "title": "Hamming Distance", "slug": "hamming-distance", "difficulty": "Easy"},
            {"id": "1318", "title": "Minimum Flips to Make a OR b Equal to c", "slug": "minimum-flips-to-make-a-or-b-equal-to-c", "difficulty": "Medium"},
            {"id": "201", "title": "Bitwise AND of Numbers Range", "slug": "bitwise-and-of-numbers-range", "difficulty": "Medium"},
            {"id": "137", "title": "Single Number II", "slug": "single-number-ii", "difficulty": "Medium"},
            {"id": "260", "title": "Single Number III", "slug": "single-number-iii", "difficulty": "Medium"}
        ],
        "math": [
            {"id": "202", "title": "Happy Number", "slug": "happy-number", "difficulty": "Easy"},
            {"id": "66", "title": "Plus One", "slug": "plus-one", "difficulty": "Easy"},
            {"id": "172", "title": "Factorial Trailing Zeroes", "slug": "factorial-trailing-zeroes", "difficulty": "Medium"},
            {"id": "69", "title": "Sqrt(x)", "slug": "sqrtx", "difficulty": "Easy"},
            {"id": "50", "title": "Pow(x, n)", "slug": "powx-n", "difficulty": "Medium"},
            {"id": "149", "title": "Max Points on a Line", "slug": "max-points-on-a-line", "difficulty": "Hard"},
            {"id": "43", "title": "Multiply Strings", "slug": "multiply-strings", "difficulty": "Medium"},
            {"id": "204", "title": "Count Primes", "slug": "count-primes", "difficulty": "Medium"},
            {"id": "48", "title": "Rotate Image", "slug": "rotate-image", "difficulty": "Medium"},
            {"id": "54", "title": "Spiral Matrix", "slug": "spiral-matrix", "difficulty": "Medium"},
            {"id": "73", "title": "Set Matrix Zeroes", "slug": "set-matrix-zeroes", "difficulty": "Medium"},
            {"id": "9", "title": "Palindrome Number", "slug": "palindrome-number", "difficulty": "Easy"},
            {"id": "12", "title": "Integer to Roman", "slug": "integer-to-roman", "difficulty": "Medium"},
            {"id": "13", "title": "Roman to Integer", "slug": "roman-to-integer", "difficulty": "Easy"},
            {"id": "1523", "title": "Count Odd Numbers in an Interval Range", "slug": "count-odd-numbers-in-an-interval-range", "difficulty": "Easy"}
        ],
        "dynamic-programming": [
            {"id": "70", "title": "Climbing Stairs", "slug": "climbing-stairs", "difficulty": "Easy"},
            {"id": "746", "title": "Min Cost Climbing Stairs", "slug": "min-cost-climbing-stairs", "difficulty": "Easy"},
            {"id": "198", "title": "House Robber", "slug": "house-robber", "difficulty": "Medium"},
            {"id": "213", "title": "House Robber II", "slug": "house-robber-ii", "difficulty": "Medium"},
            {"id": "5", "title": "Longest Palindromic Substring", "slug": "longest-palindromic-substring", "difficulty": "Medium"},
            {"id": "647", "title": "Palindromic Substrings", "slug": "palindromic-substrings", "difficulty": "Medium"},
            {"id": "91", "title": "Decode Ways", "slug": "decode-ways", "difficulty": "Medium"},
            {"id": "322", "title": "Coin Change", "slug": "coin-change", "difficulty": "Medium"},
            {"id": "152", "title": "Maximum Product Subarray", "slug": "maximum-product-subarray", "difficulty": "Medium"},
            {"id": "139", "title": "Word Break", "slug": "word-break", "difficulty": "Medium"},
            {"id": "300", "title": "Longest Increasing Subsequence", "slug": "longest-increasing-subsequence", "difficulty": "Medium"},
            {"id": "416", "title": "Partition Equal Subset Sum", "slug": "partition-equal-subset-sum", "difficulty": "Medium"},
            {"id": "62", "title": "Unique Paths", "slug": "unique-paths", "difficulty": "Medium"},
            {"id": "1143", "title": "Longest Common Subsequence", "slug": "longest-common-subsequence", "difficulty": "Medium"},
            {"id": "309", "title": "Best Time to Buy and Sell Stock with Cooldown", "slug": "best-time-to-buy-and-sell-stock-with-cooldown", "difficulty": "Medium"}
        ],
        "graph": [
            {"id": "200", "title": "Number of Islands", "slug": "number-of-islands", "difficulty": "Medium"},
            {"id": "695", "title": "Max Area of Island", "slug": "max-area-of-island", "difficulty": "Medium"},
            {"id": "133", "title": "Clone Graph", "slug": "clone-graph", "difficulty": "Medium"},
            {"id": "286", "title": "Walls and Gates", "slug": "walls-and-gates", "difficulty": "Medium"},
            {"id": "994", "title": "Rotting Oranges", "slug": "rotting-oranges", "difficulty": "Medium"},
            {"id": "417", "title": "Pacific Atlantic Water Flow", "slug": "pacific-atlantic-water-flow", "difficulty": "Medium"},
            {"id": "130", "title": "Surrounded Regions", "slug": "surrounded-regions", "difficulty": "Medium"},
            {"id": "207", "title": "Course Schedule", "slug": "course-schedule", "difficulty": "Medium"},
            {"id": "210", "title": "Course Schedule II", "slug": "course-schedule-ii", "difficulty": "Medium"},
            {"id": "261", "title": "Graph Valid Tree", "slug": "graph-valid-tree", "difficulty": "Medium"},
            {"id": "323", "title": "Number of Connected Components in an Undirected Graph", "slug": "number-of-connected-components-in-an-undirected-graph", "difficulty": "Medium"},
            {"id": "684", "title": "Redundant Connection", "slug": "redundant-connection", "difficulty": "Medium"},
            {"id": "127", "title": "Word Ladder", "slug": "word-ladder", "difficulty": "Hard"},
            {"id": "743", "title": "Network Delay Time", "slug": "network-delay-time", "difficulty": "Medium"},
            {"id": "787", "title": "Cheapest Flights Within K Stops", "slug": "cheapest-flights-within-k-stops", "difficulty": "Medium"}
        ]
    }
    
    # 3. Fetch roadmap data (Try DB first, then Fallback)
    db_data = memory.get_roadmap_data()
    
    # Check if DB has valid data (at least one topic with problems)
    has_valid_db_data = False
    if db_data and db_data.get("roadmap"):
        for t in db_data["roadmap"]:
            if t.get("problems") and len(t["problems"]) > 0:
                has_valid_db_data = True
                break
    
    if has_valid_db_data:
        print("Using roadmap data from Database.")
        roadmap_data = db_data["roadmap"]
    else:
        # If DB is empty or invalid, use fallback data and store it
        print("Database empty or incomplete. Initializing roadmap data from Fallback...")
        roadmap_data = []
        
        # Prepare data for storage
        problems_by_tag = FALLBACK_PROBLEMS
        
        # Store in DB
        try:
            memory.store_roadmap_data(topics, problems_by_tag)
            print("Stored roadmap data in DB.")
        except Exception as e:
            print(f"Error storing roadmap data: {e}")
        
        # Construct response structure from fallback
        for topic in topics:
            problems = []
            fallback_list = FALLBACK_PROBLEMS.get(topic["tag"], [])
            for q in fallback_list:
                problems.append({
                    "id": q["id"],
                    "title": q["title"],
                    "slug": q["slug"],
                    "difficulty": q["difficulty"],
                    "url": f"https://leetcode.com/problems/{q['slug']}/",
                    "solved": False
                })
            
            roadmap_data.append({
                "topicId": topic["id"],
                "title": topic["name"],
                "tag": topic["tag"],
                "totalProblems": len(problems),
                "solvedProblems": 0,
                "problems": problems
            })

    # 4. Update with User Progress (Solved Status)
    for topic_data in roadmap_data:
        # Try to fetch live data if needed, but primarily rely on DB + Solved Status
        # (Optional: You could still try to fetch new problems from LeetCode here and update DB)
        
        problems = topic_data["problems"]
        solved_count = 0
        
        for p in problems:
            is_solved = p["slug"] in solved_slugs
            p["solved"] = is_solved
            if is_solved:
                solved_count += 1
        
        topic_data["solvedProblems"] = solved_count
        topic_data["totalProblems"] = len(problems)
        
    return jsonify({
        "success": True,
        "username": username,
        "roadmap": roadmap_data
    })

@app.route("/api/weakness", methods=["GET"])
def get_weakness_api():
    """Get weakness analysis"""
    try:
        return jsonify({"success": True, "weaknesses": []})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- Training / Resources API ---

@app.route('/api/training/resources', methods=['GET'])
def get_training_resources():
    from maang_agent.memory_persistence import AgentMemoryManager
    memory = AgentMemoryManager()
    
    # Get progress from DB
    progress_list = memory.get_training_progress()
    progress_map = {p['resource_id']: p for p in progress_list}
    
    resources = []
    
    # 1. DSA (Notion)
    dsa_resource = {
        "id": "dsa_notion",
        "title": "Data Structures and Algorithms (Notion)",
        "type": "dsa",
        "url": "https://www.notion.so/Data-Structures-and-Algorithms-2b168fc54025814592a3f9267989f9b1?source=copy_link",
        "status": progress_map.get("dsa_notion", {}).get("status", "Not Started")
    }
    resources.append(dsa_resource)
    
    # 2. System Design (userData files - filtered list)
    allowed_books = [
        "Cracking-the-Coding-Interview-6th-Edition-189-Programming-Questions-and-Solutions.pdf",
        "competitive.programming.3rd.edition.pdf",
        "Designing Data Intensive Applications by Martin Kleppmann.pdf",
        "System Design Interview by Alex Xu.pdf",
        "System-Design-Alex-Xu-Vol-2.pdf"
    ]
    
    user_data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "userData")
    if os.path.exists(user_data_dir):
        for filename in os.listdir(user_data_dir):
            if filename in allowed_books:
                res_id = f"sys_design_{filename}"
                resources.append({
                    "id": res_id,
                    "title": filename.replace('.pdf', '').replace('-', ' '),
                    "type": "system_design",
                    "filename": filename,
                    "url": f"http://localhost:5100/userdata/{filename}",
                    "status": progress_map.get(res_id, {}).get("status", "Not Started")
                })
    
    return jsonify({"success": True, "resources": resources})


# ===== AUTHENTICATION ENDPOINTS =====

@app.route('/api/auth/signup', methods=['POST'])
def signup():
    """User registration"""
    data = request.json
    email = data.get('email')
    password = data.get('password')
    full_name = data.get('full_name')
    username = data.get('username', email.split('@')[0])  # Use email prefix as default username
    
    if not email or not password:
        return jsonify({"success": False, "error": "Email and password required"}), 400
    
    # create_user(username, email, password, full_name)
    user_id = create_user(username, email, password, full_name)
    if not user_id:
        return jsonify({"success": False, "error": "User already exists"}), 409
    
    # Generate token
    token = generate_token(user_id, email)
    
    return jsonify({
        "success": True,
        "token": token,
        "user": {"id": user_id, "email": email, "full_name": full_name}
    })

@app.route('/api/auth/login', methods=['POST'])
def login():
    """User login"""
    data = request.json
    email = data.get('email')
    password = data.get('password')
    
    if not email or not password:
        return jsonify({"success": False, "error": "Email and password required"}), 400
    
    user = authenticate_user(email, password)
    if not user:
        return jsonify({"success": False, "error": "Invalid credentials"}), 401
    
    # Generate token
    token = generate_token(user['id'], user['email'])
    
    return jsonify({
        "success": True,
        "token": token,
        "user": user
    })

@app.route('/api/auth/me', methods=['GET'])
@require_auth
def get_current_user():
    """Get current user info"""
    user = get_user_by_id(request.user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    
    return jsonify({"success": True, "user": user})

# ===== PLATFORM CREDENTIALS =====

@app.route('/api/credentials/leetcode', methods=['POST'])
@require_auth
def save_leetcode_credentials():
    """Save LeetCode credentials and authenticate"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400
    
    # Authenticate with LeetCode
    result = call_mcp("leetcode_login", {"username": username, "password": password})
    
    if not result.get("ok"):
        return jsonify({"success": False, "error": result.get("error", "Authentication failed")}), 401
    
    session = result.get("session")
    
    # Save credentials
    save_user_credentials(
        user_id=request.user_id,
        platform='leetcode',
        username=username,
        session_cookie=session
    )
    
    # Invalidate cache
    CacheManager.invalidate_user(request.user_id)
    
    return jsonify({"success": True, "message": "LeetCode credentials saved successfully"})

@app.route('/api/credentials/github', methods=['POST'])
@require_auth
def save_github_credentials():
    """Save GitHub credentials"""
    data = request.json
    username = data.get('username')
    token = data.get('token')
    
    if not username:
        return jsonify({"success": False, "error": "Username required"}), 400
    
    save_user_credentials(
        user_id=request.user_id,
        platform='github',
        username=username,
        encrypted_token=token
    )
    
    CacheManager.invalidate_user(request.user_id)
    
    return jsonify({"success": True, "message": "GitHub credentials saved successfully"})

# ===== OPTIMIZED SYNC ENDPOINTS =====

@app.route('/api/sync/leetcode', methods=['POST'])
@require_auth
def sync_leetcode():
    """Sync LeetCode data with intelligent caching and weakness analysis"""
    try:
        force_refresh = request.json.get('force_refresh', False) if request.json else False
        
        # Run async sync
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(SyncService.sync_leetcode_data(request.user_id, force_refresh))
        
        # Analyze weaknesses if sync succeeded
        if result.get('success'):
            weaknesses = loop.run_until_complete(SyncService.analyze_weaknesses(request.user_id, result))
            result['weaknesses_analyzed'] = len(weaknesses)
        
        loop.close()
        
        if result.get('success'):
            return jsonify(result)
        else:
            app.logger.error(f"LeetCode sync failed: {result.get('error')}")
            return jsonify(result), 500
    except Exception as e:
        app.logger.error(f"LeetCode sync exception: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/sync/github', methods=['POST'])
@require_auth
def sync_github():
    """Sync GitHub data with intelligent caching"""
    force_refresh = request.json.get('force_refresh', False) if request.json else False
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    result = loop.run_until_complete(SyncService.sync_github_data(request.user_id, force_refresh))
    loop.close()
    
    if result.get('success'):
        return jsonify(result)
    else:
        return jsonify(result), 500

@app.route('/api/sync/all', methods=['POST'])
@require_auth
def sync_all():
    """
    Full sync of all platforms with parallel processing
    This is the innovative refresh button endpoint
    """
    force_refresh = request.json.get('force_refresh', False) if request.json else False
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    result = loop.run_until_complete(SyncService.full_sync(request.user_id, force_refresh))
    loop.close()
    
    return jsonify({
        "success": True,
        "data": result,
        "message": "All platforms synced successfully"
    })

@app.route('/api/weaknesses', methods=['GET'])
@require_auth
def api_get_weaknesses():
    """Get AI-analyzed weaknesses for current user"""
    from memory.db import get_user_weaknesses
    weaknesses = get_user_weaknesses(request.user_id)
    return jsonify({"success": True, "weaknesses": weaknesses})

@app.route('/api/user/progress', methods=['GET'])
@require_auth
def api_get_user_progress():
    """Get comprehensive user progress"""
    try:
        from services.sync_service import SyncService
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        progress = loop.run_until_complete(SyncService.get_comprehensive_progress(request.user_id))
        loop.close()
        
        return jsonify({"success": True, "data": progress})
    except Exception as e:
        app.logger.error(f"Progress fetch error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# ===== CACHE MANAGEMENT =====

@app.route('/api/cache/clear', methods=['POST'])
@require_auth
def clear_user_cache():
    """Clear all cache for current user"""
    CacheManager.invalidate_user(request.user_id)
    return jsonify({"success": True, "message": "Cache cleared successfully"})

@app.route('/api/user/focus', methods=['GET', 'POST'])
@require_auth
def user_focus():
    """Get or set user's current focus topic"""
    from memory.db import save_user_focus, get_user_focus
    
    if request.method == 'POST':
        data = request.json
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400
            
        success = save_user_focus(request.user_id, data)
        if success:
            return jsonify({"success": True, "message": "Focus topic updated"})
        else:
            return jsonify({"success": False, "error": "Failed to update focus topic"}), 500
            
    else: # GET
        focus = get_user_focus(request.user_id)
        return jsonify({"success": True, "focus": focus})

@app.route('/api/training/progress', methods=['POST'])
def update_training_progress():
    data = request.json
    resource_id = data.get('resource_id')
    title = data.get('title')
    res_type = data.get('type')
    status = data.get('status')
    
    from maang_agent.memory_persistence import AgentMemoryManager
    memory = AgentMemoryManager()
    memory.update_training_progress(resource_id, title, res_type, status)
    
    return jsonify({"success": True})

@app.route('/userdata/<path:filename>')
def serve_userdata(filename):
    user_data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "userData")
    return send_from_directory(user_data_dir, filename)

@app.route('/api/training/notes', methods=['POST'])
def save_note():
    data = request.json
    from maang_agent.memory_persistence import AgentMemoryManager
    memory = AgentMemoryManager()
    note_id = memory.save_reading_note(
        data.get('resource_id'),
        data.get('page_number', 0),
        data.get('note_type', 'note'),
        data.get('content', '')
    )
    return jsonify({"success": True, "note_id": note_id})

@app.route('/api/training/notes/<resource_id>', methods=['GET'])
def get_notes(resource_id):
    from maang_agent.memory_persistence import AgentMemoryManager
    memory = AgentMemoryManager()
    notes = memory.get_reading_notes(resource_id)
    return jsonify({"success": True, "notes": notes})

@app.route('/api/leetcode-auth', methods=['POST'])
def leetcode_auth():
    """Authenticate with LeetCode and store session"""
    from memory.db import set_setting
    from tracker.tracker import call_mcp
    
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400
    
    # Call MCP login tool
    result = call_mcp("leetcode_login", {"username": username, "password": password})
    
    if result.get("ok"):
        session = result.get("session")
        # Store the session and username
        set_setting('leetcode_user', username)
        set_setting('leetcode_session', session)
        return jsonify({"success": True, "message": "Successfully authenticated with LeetCode"})
    else:
        return jsonify({"success": False, "error": result.get("error", "Authentication failed")}), 401

@app.route('/api/settings', methods=['GET', 'POST'])
def handle_settings():
    from memory.db import get_setting, set_setting
    if request.method == 'POST':
        data = request.json
        if 'leetcode_user' in data: set_setting('leetcode_user', data['leetcode_user'])
        if 'github_user' in data: set_setting('github_user', data['github_user'])
        if 'leetcode_session' in data: set_setting('leetcode_session', data['leetcode_session'])
        if 'github_token' in data: set_setting('github_token', data['github_token'])
        return jsonify({"success": True})
    else:
        return jsonify({
            "leetcode_user": get_setting('leetcode_user', os.getenv("LEETCODE_USERNAME", "")),
            "github_user": get_setting('github_user', os.getenv("GITHUB_USERNAME", "")),
            # Return empty string for secrets to avoid exposing them, but indicate if set
            "leetcode_session_set": bool(get_setting('leetcode_session')),
            "github_token_set": bool(get_setting('github_token'))
        })

@app.route('/api/sync-stats', methods=['POST', 'GET'])
def sync_stats():
    from memory.db import get_setting
    # Prioritize DB setting, fallback to env, then default
    lc_user = get_setting('leetcode_user', os.getenv("LEETCODE_USERNAME", "arnold-1324"))
    gh_user = get_setting('github_user', os.getenv("GITHUB_USERNAME", "arnold-1324"))
    
    lc_session = get_setting('leetcode_session')
    gh_token = get_setting('github_token')
    
    try:
        # Use the imported snapshot functions with auth
        lc_res = snapshot_leetcode(lc_user, lc_session)
        gh_res = snapshot_github(gh_user, gh_token)
        return jsonify({
            "success": True,
            "leetcode": lc_res,
            "github": gh_res
        })
    except Exception as e:
        print(f"Error syncing stats: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/export/excel', methods=['GET'])
def export_to_excel():
    """Export all database tables to Excel format"""
    import sqlite3
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file
    
    try:
        DB_PATH = os.getenv("SQLITE_PATH", "./memory.db")
        
        # Check if database exists
        if not os.path.exists(DB_PATH):
            return jsonify({"success": False, "error": "Database not found"}), 404
        
        # Connect to database
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Get all tables
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row[0] for row in cursor.fetchall()]
        
        if not tables:
            conn.close()
            return jsonify({"success": False, "error": "No tables found"}), 404
        
        # Create workbook
        workbook = Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        total_records = 0
        
        # Export each table
        for table_name in tables:
            try:
                # Get table data
                cursor.execute(f"SELECT * FROM {table_name}")
                rows = cursor.fetchall()
                
                # Get column names
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns = [col[1] for col in cursor.fetchall()]
                
                # Create sheet (limit name to 31 chars for Excel)
                sheet = workbook.create_sheet(title=table_name[:31])
                
                # Write headers
                for col_idx, column_name in enumerate(columns, 1):
                    cell = sheet.cell(row=1, column=col_idx, value=column_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Write data
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Auto-adjust column widths
                for col_idx, column_name in enumerate(columns, 1):
                    max_length = len(column_name)
                    for row_idx in range(2, min(len(rows) + 2, 100)):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                
                # Freeze header row
                sheet.freeze_panes = "A2"
                
                total_records += len(rows)
                
            except Exception as e:
                print(f"Error exporting table {table_name}: {e}")
                continue
        
        # Create summary sheet
        summary_sheet = workbook.create_sheet(title="📊 Summary", index=0)
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 15
        
        summary_sheet['A1'] = "MAANG Tracker - Database Export Summary"
        summary_sheet['A1'].font = Font(bold=True, size=14, color="4472C4")
        summary_sheet.merge_cells('A1:B1')
        
        summary_sheet['A3'] = "Export Date:"
        summary_sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary_sheet['A4'] = "Database Path:"
        summary_sheet['B4'] = DB_PATH
        summary_sheet['A5'] = "Total Tables:"
        summary_sheet['B5'] = len(tables)
        summary_sheet['A6'] = "Total Records:"
        summary_sheet['B6'] = total_records
        
        summary_sheet['A8'] = "Table Name"
        summary_sheet['B8'] = "Record Count"
        summary_sheet['A8'].font = Font(bold=True)
        summary_sheet['B8'].font = Font(bold=True)
        
        row = 9
        for table_name in tables:
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            count = cursor.fetchone()[0]
            summary_sheet[f'A{row}'] = table_name
            summary_sheet[f'B{row}'] = count
            row += 1
        
        conn.close()
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"maang_tracker_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

# ===== LEETCODE SESSION MANAGEMENT =====

@app.route("/api/leetcode/session", methods=["POST"])
def save_leetcode_session():
    """Save LeetCode session credentials"""
    try:
        from services.leetcode_session import LeetCodeSessionManager
        
        data = request.json
        username = data.get("username")
        session_id = data.get("session_id")
        
        if not username or not session_id:
            return jsonify({"success": False, "error": "Username and session ID required"}), 400
        
        manager = LeetCodeSessionManager()
        manager.save_session(username, session_id)
        
        return jsonify({
            "success": True,
            "message": "LeetCode session saved successfully"
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/leetcode/session", methods=["GET"])
def get_leetcode_session():
    """Get saved LeetCode session (without exposing session_id)"""
    try:
        from services.leetcode_session import LeetCodeSessionManager
        
        manager = LeetCodeSessionManager()
        session = manager.get_session()
        
        if session:
            return jsonify({
                "success": True,
                "data": {
                    "username": session.get("username"),
                    "has_session": True,
                    "saved_at": session.get("saved_at")
                }
            })
        else:
            return jsonify({
                "success": True,
                "data": {"has_session": False}
            })
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/leetcode/session", methods=["DELETE"])
def delete_leetcode_session():
    """Clear saved LeetCode session"""
    try:
        from services.leetcode_session import LeetCodeSessionManager
        
        manager = LeetCodeSessionManager()
        manager.clear_session()
        
        return jsonify({
            "success": True,
            "message": "Session cleared successfully"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/leetcode/stats", methods=["GET"])
def get_leetcode_stats():
    """Fetch LeetCode statistics using saved session"""
    try:
        from services.leetcode_session import LeetCodeSessionManager, fetch_leetcode_stats
        
        manager = LeetCodeSessionManager()
        session = manager.get_session()
        
        if not session:
            return jsonify({
                "success": False,
                "error": "No LeetCode session found. Please configure your session first."
            }), 400
        
        username = session.get("username")
        session_id = session.get("session_id")
        
        result = fetch_leetcode_stats(username, session_id)
        
        if result.get("success"):
            return jsonify(result)
        else:
            return jsonify(result), 400
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/leetcode/test-session", methods=["POST"])
def test_leetcode_session():
    """Test if provided session credentials are valid"""
    try:
        from services.leetcode_session import fetch_leetcode_stats
        
        data = request.json
        username = data.get("username")
        session_id = data.get("session_id")
        
        if not username or not session_id:
            return jsonify({"success": False, "error": "Username and session ID required"}), 400
        
        result = fetch_leetcode_stats(username, session_id)
        
        if result.get("success"):
            return jsonify({
                "success": True,
                "message": "Session is valid!",
                "stats": result.get("data")
            })
        else:
            return jsonify({
                "success": False,
                "error": result.get("error", "Invalid session")
            }), 400
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


if __name__ == "__main__":
    # Use SocketIO if available, otherwise use regular Flask app
    if socketio_instance:
        socketio_instance.run(app, host="0.0.0.0", port=5100, debug=True)
    else:
        app.run(host="0.0.0.0", port=5100, debug=True)
