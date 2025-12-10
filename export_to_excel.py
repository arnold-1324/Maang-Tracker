"""
Export all database tables to Excel format
Creates a multi-sheet Excel file with all data from the database
"""
import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Database path
DB_PATH = os.getenv("SQLITE_PATH", "./memory.db")

def get_all_tables():
    """Get list of all tables in the database"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [row[0] for row in cursor.fetchall()]
    conn.close()
    return tables

def export_table_to_sheet(conn, table_name, workbook):
    """Export a single table to an Excel sheet"""
    cursor = conn.cursor()
    
    # Get table data
    cursor.execute(f"SELECT * FROM {table_name}")
    rows = cursor.fetchall()
    
    # Get column names
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = [col[1] for col in cursor.fetchall()]
    
    # Create sheet
    sheet = workbook.create_sheet(title=table_name[:31])  # Excel sheet name limit is 31 chars
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, column_name in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=col_idx, value=column_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for col_idx, column_name in enumerate(columns, 1):
        max_length = len(column_name)
        for row_idx in range(2, min(len(rows) + 2, 100)):  # Check first 100 rows for performance
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width (max 50 characters)
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Freeze header row
    sheet.freeze_panes = "A2"
    
    return len(rows)

def export_all_to_excel(output_filename=None):
    """Export all database tables to Excel"""
    
    if output_filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"maang_tracker_export_{timestamp}.xlsx"
    
    # Check if database exists
    if not os.path.exists(DB_PATH):
        print(f"❌ Database not found at: {DB_PATH}")
        return None
    
    print(f"📊 Starting export from database: {DB_PATH}")
    print(f"📁 Output file: {output_filename}\n")
    
    # Connect to database
    conn = sqlite3.connect(DB_PATH)
    
    # Create workbook
    workbook = Workbook()
    # Remove default sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    
    # Get all tables
    tables = get_all_tables()
    
    if not tables:
        print("❌ No tables found in database")
        conn.close()
        return None
    
    print(f"Found {len(tables)} tables to export:\n")
    
    # Export each table
    total_records = 0
    for table_name in tables:
        try:
            record_count = export_table_to_sheet(conn, table_name, workbook)
            total_records += record_count
            print(f"✅ {table_name:<30} - {record_count:>6} records")
        except Exception as e:
            print(f"❌ {table_name:<30} - Error: {str(e)}")
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet(title="📊 Summary", index=0)
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    
    # Summary header
    summary_sheet['A1'] = "MAANG Tracker - Database Export Summary"
    summary_sheet['A1'].font = Font(bold=True, size=14, color="4472C4")
    summary_sheet.merge_cells('A1:B1')
    
    summary_sheet['A3'] = "Export Date:"
    summary_sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_sheet['A4'] = "Database Path:"
    summary_sheet['B4'] = DB_PATH
    summary_sheet['A5'] = "Total Tables:"
    summary_sheet['B5'] = len(tables)
    summary_sheet['A6'] = "Total Records:"
    summary_sheet['B6'] = total_records
    
    summary_sheet['A8'] = "Table Name"
    summary_sheet['B8'] = "Record Count"
    summary_sheet['A8'].font = Font(bold=True)
    summary_sheet['B8'].font = Font(bold=True)
    
    row = 9
    for table_name in tables:
        cursor = conn.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
        count = cursor.fetchone()[0]
        summary_sheet[f'A{row}'] = table_name
        summary_sheet[f'B{row}'] = count
        row += 1
    
    # Close database connection
    conn.close()
    
    # Save workbook
    workbook.save(output_filename)
    
    print(f"\n{'='*60}")
    print(f"✅ Export completed successfully!")
    print(f"📁 File saved: {os.path.abspath(output_filename)}")
    print(f"📊 Total records exported: {total_records}")
    print(f"{'='*60}\n")
    
    return output_filename

if __name__ == "__main__":
    # You can specify a custom filename or let it auto-generate with timestamp
    export_all_to_excel()
