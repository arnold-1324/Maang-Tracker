# analyzer/user_data_analyzer.py
"""
Analyze user data files (Excel trackers, problem lists, etc.)
to extract and classify problems for custom weakness tuning.
"""

import os
import json
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path

class UserDataAnalyzer:
    """Parse and analyze user data files for problem classification."""
    
    def __init__(self, data_dir: str = "./userData"):
        self.data_dir = Path(data_dir)
        self.problems = []
        
    def analyze_excel_tracker(self, filename: str = "Maang tracker (1).xlsx") -> List[Dict]:
        """Parse MAANG tracker Excel file and extract problems."""
        filepath = self.data_dir / filename
        
        if not filepath.exists():
            print(f"⚠️ File not found: {filepath}")
            return []
        
        problems = []
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(filepath)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Assume first row is header
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    problem = self._parse_tracker_row(headers, row, sheet_name)
                    if problem:
                        problems.append(problem)
            
            print(f"✅ Parsed {len(problems)} problems from {filename}")
            self.problems.extend(problems)
            return problems
            
        except ImportError:
            print("⚠️ openpyxl not installed. Install with: pip install openpyxl")
            return []
        except Exception as e:
            print(f"❌ Error parsing Excel: {e}")
            return []
    
    def _parse_tracker_row(self, headers: List[str], row: tuple, 
                          sheet_name: str) -> Optional[Dict]:
        """Parse a row from the tracker and extract problem info."""
        data = dict(zip(headers, row))
        
        # Common column mappings (adjust based on actual tracker format)
        problem = {
            "source": sheet_name.lower() if sheet_name else "unknown",
            "title": str(data.get("Problem") or data.get("Title") or ""),
            "difficulty": str(data.get("Difficulty") or "Medium").lower(),
            "status": str(data.get("Status") or data.get("Solved") or ""),
            "tags": self._extract_tags(data.get("Tags") or ""),
            "date_solved": str(data.get("Date Solved") or data.get("Date") or ""),
            "notes": str(data.get("Notes") or ""),
        }
        
        if not problem["title"]:
            return None
        
        return problem
    
    def _extract_tags(self, tags_str: str) -> List[str]:
        """Parse tag string into list."""
        if not tags_str:
            return []
        
        # Handle comma or semicolon separated tags
        separator = "," if "," in tags_str else ";"
        return [tag.strip().lower() for tag in tags_str.split(separator) if tag.strip()]
    
    def extract_pdf_topics(self) -> Dict[str, List[str]]:
        """Extract chapter/section titles from reference PDFs."""
        topics = {
            "algorithms": [],
            "data_structures": [],
            "system_design": [],
            "interview_prep": []
        }
        
        pdf_mappings = {
            "Cracking-the-Coding-Interview-6th-Edition-189-Programming-Questions-and-Solutions.pdf": "interview_prep",
            "Competitive.programming.3rd.edition.pdf": "algorithms",
            "Designing Data Intensive Applications by Martin Kleppmann.pdf": "system_design",
            "System Design Interview by Alex Xu.pdf": "system_design"
        }
        
        try:
            import PyPDF2
            
            for pdf_file, category in pdf_mappings.items():
                filepath = self.data_dir / pdf_file
                
                if not filepath.exists():
                    continue
                
                try:
                    with open(filepath, 'rb') as f:
                        reader = PyPDF2.PdfReader(f)
                        # Extract text from first page for quick topic extraction
                        if reader.pages:
                            text = reader.pages[0].extract_text()
                            # Basic chapter extraction (adjust based on PDF structure)
                            chapters = [line.strip() for line in text.split('\n') 
                                      if line.strip() and any(
                                          keyword in line.lower() 
                                          for keyword in ['chapter', 'section', 'module']
                                      )]
                            topics[category].extend(chapters[:5])  # First 5 chapters
                
                except Exception as e:
                    print(f"⚠️ Could not process {pdf_file}: {e}")
        
        except ImportError:
            print("⚠️ PyPDF2 not installed. Install with: pip install PyPDF2")
        
        return topics
    
    def generate_weakness_profile(self) -> Dict[str, int]:
        """Generate weakness profile from user data."""
        weakness_profile = {}
        
        for problem in self.problems:
            # Weight by difficulty and tags
            difficulty_weight = {
                "easy": 1,
                "medium": 3,
                "hard": 5
            }.get(problem.get("difficulty", "medium"), 2)
            
            # Lower score for solved problems, higher for unsolved
            status = (problem.get("status") or "").lower()
            if "solved" in status or "ac" in status:
                score_modifier = 0.3  # Solved = less weakness
            elif "attempted" in status or "in progress" in status:
                score_modifier = 0.7  # Attempted = medium weakness
            else:
                score_modifier = 1.0  # Unsolved = high weakness
            
            final_score = int(difficulty_weight * score_modifier * 20)
            
            # Distribute score across tags
            tags = problem.get("tags", [])
            if not tags:
                tags = ["general"]
            
            for tag in tags:
                if tag not in weakness_profile:
                    weakness_profile[tag] = 0
                weakness_profile[tag] += final_score
        
        # Normalize scores (0-100)
        if weakness_profile:
            max_score = max(weakness_profile.values())
            for key in weakness_profile:
                weakness_profile[key] = min(100, int((weakness_profile[key] / max_score) * 100))
        
        return weakness_profile
    
    def recommend_study_plan(self, weakness_profile: Dict[str, int]) -> List[Dict]:
        """Generate personalized study plan based on weaknesses."""
        recommendations = []
        
        # Sort by weakness score (descending)
        sorted_weaknesses = sorted(
            weakness_profile.items(), 
            key=lambda x: x[1], 
            reverse=True
        )
        
        for topic, score in sorted_weaknesses[:5]:  # Top 5 weaknesses
            intensity = "high" if score > 70 else "medium" if score > 40 else "low"
            
            recommendation = {
                "topic": topic,
                "weakness_score": score,
                "intensity": intensity,
                "estimated_hours": {
                    "high": 20,
                    "medium": 10,
                    "low": 5
                }[intensity],
                "resources": self._get_topic_resources(topic),
                "priority": 1 if score > 80 else 2 if score > 60 else 3
            }
            
            recommendations.append(recommendation)
        
        return recommendations
    
    def _get_topic_resources(self, topic: str) -> List[Dict]:
        """Get recommended resources for a topic."""
        resource_map = {
            "array": [
                {"title": "LeetCode Arrays", "url": "https://leetcode.com/tag/array/", "source": "leetcode"},
                {"title": "GeeksforGeeks Arrays", "url": "https://www.geeksforgeeks.org/array-data-structure/", "source": "gfg"},
                {"title": "TakeUForward Arrays", "url": "https://takeuforward.org/plus/dsa/", "source": "takeuforward"}
            ],
            "dynamic-programming": [
                {"title": "LeetCode DP", "url": "https://leetcode.com/tag/dynamic-programming/", "source": "leetcode"},
                {"title": "GeeksforGeeks DP", "url": "https://www.geeksforgeeks.org/dynamic-programming/", "source": "gfg"},
                {"title": "Competitive Programming DP", "url": "https://takeuforward.org/plus/dsa/", "source": "takeuforward"}
            ],
            "graph": [
                {"title": "LeetCode Graph", "url": "https://leetcode.com/tag/graph/", "source": "leetcode"},
                {"title": "GeeksforGeeks Graph", "url": "https://www.geeksforgeeks.org/graph-data-structure-and-algorithms/", "source": "gfg"}
            ],
            "system-design": [
                {"title": "System Design Interview", "url": "https://takeuforward.org/plus/dsa/", "source": "takeuforward"},
                {"title": "Alex Xu's Resources", "url": "https://www.youtube.com/c/AlexXu", "source": "external"}
            ]
        }
        
        # Default resources if topic not in map
        default_resources = [
            {"title": f"LeetCode {topic}", "url": f"https://leetcode.com/tag/{topic}/", "source": "leetcode"},
            {"title": f"GeeksforGeeks {topic}", "url": "https://www.geeksforgeeks.org/", "source": "gfg"}
        ]
        
        return resource_map.get(topic.lower(), default_resources)
    
    def export_analysis(self, output_file: str = "analysis_report.json") -> str:
        """Export analysis results to JSON."""
        weakness_profile = self.generate_weakness_profile()
        recommendations = self.recommend_study_plan(weakness_profile)
        pdf_topics = self.extract_pdf_topics()
        
        report = {
            "generated_at": datetime.now().isoformat(),
            "problems_analyzed": len(self.problems),
            "weakness_profile": weakness_profile,
            "recommendations": recommendations,
            "pdf_topics": pdf_topics,
            "problems": self.problems[:100]  # First 100 for reference
        }
        
        output_path = Path(output_file)
        with open(output_path, 'w') as f:
            json.dump(report, f, indent=2)
        
        print(f"✅ Analysis exported to {output_path}")
        return str(output_path)


# CLI for standalone analysis
if __name__ == "__main__":
    analyzer = UserDataAnalyzer()
    
    print("📊 Analyzing user data files...")
    analyzer.analyze_excel_tracker()
    
    print("\n📚 Extracting PDF topics...")
    pdf_topics = analyzer.extract_pdf_topics()
    print(f"Found topics in PDFs: {pdf_topics}")
    
    print("\n💪 Generating weakness profile...")
    weakness_profile = analyzer.generate_weakness_profile()
    print(f"Weakness Profile: {weakness_profile}")
    
    print("\n🎯 Generating study plan...")
    recommendations = analyzer.recommend_study_plan(weakness_profile)
    for rec in recommendations:
        print(f"  • {rec['topic']}: {rec['weakness_score']}% weakness (Priority: {rec['priority']})")
    
    print("\n📤 Exporting analysis...")
    analyzer.export_analysis()
